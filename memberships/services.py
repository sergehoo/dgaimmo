"""Services métier pour l'app `memberships`.

- ``parse_members_file()`` : import en masse de membres depuis un fichier
  Excel (.xlsx) ou CSV (.csv). Validation par ligne, retour d'un rapport
  détaillé (membres créés + erreurs).
- ``send_member_invitation()`` : envoi d'un mail d'invitation contenant
  un lien d'auto-onboarding.
"""

from __future__ import annotations

import csv
import io
import uuid
from dataclasses import dataclass, field
from datetime import timedelta
from typing import Iterable

from django.conf import settings
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils import timezone

from memberships.models import Member, MemberInvitation


# --- Mapping colonnes Excel → champs Member ---------------------------------
# Toléré : nom de colonne en français, accents et casse ignorés.
COLUMN_ALIASES = {
    "nom": "last_name",
    "lastname": "last_name",
    "last_name": "last_name",
    "prenom": "first_name",
    "prenoms": "first_name",
    "firstname": "first_name",
    "first_name": "first_name",
    "telephone": "phone",
    "phone": "phone",
    "email": "email",
    "mail": "email",
    "date_naissance": "birth_date",
    "datedenaissance": "birth_date",
    "birth_date": "birth_date",
    "genre": "gender",
    "sexe": "gender",
    "gender": "gender",
    "cni": "national_id",
    "passeport": "national_id",
    "national_id": "national_id",
    "lieu_naissance": "birth_place",
    "ville_naissance": "birth_place",
    "birth_place": "birth_place",
    "situation_matrimoniale": "marital_status",
    "marital_status": "marital_status",
    "personnes_a_charge": "dependents_count",
    "dependents_count": "dependents_count",
    "code_membre": "member_code",
    "member_code": "member_code",
}

REQUIRED_FIELDS = {"first_name", "last_name", "phone"}

# Mapping libellés humains → codes choices acceptés
MARITAL_ALIASES = {
    "celibataire": "single",
    "single": "single",
    "marie": "married",
    "mariee": "married",
    "married": "married",
    "divorce": "divorced",
    "divorcee": "divorced",
    "divorced": "divorced",
    "veuf": "widowed",
    "veuve": "widowed",
    "widowed": "widowed",
    "union_libre": "union_libre",
    "concubinage": "union_libre",
}

GENDER_ALIASES = {
    "h": "male",
    "homme": "male",
    "m": "male",
    "male": "male",
    "f": "female",
    "femme": "female",
    "female": "female",
    "autre": "other",
    "other": "other",
}


def _normalize(value: str) -> str:
    """Slugifie sommairement (lowercase, sans espaces ni accents) pour matcher les alias."""
    if not value:
        return ""
    import unicodedata

    normalized = (
        unicodedata.normalize("NFKD", str(value))
        .encode("ascii", "ignore")
        .decode("ascii")
    )
    return normalized.strip().lower().replace(" ", "_").replace("-", "_")


@dataclass
class ImportError:
    line: int
    field: str
    message: str


@dataclass
class ImportReport:
    created: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    skipped_duplicates: int = 0

    @property
    def created_count(self) -> int:
        return len(self.created)

    @property
    def error_count(self) -> int:
        return len(self.errors)


def _read_rows(file_obj, filename: str) -> Iterable[dict]:
    """Itère sur les lignes du fichier (Excel .xlsx OU CSV)."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl n'est pas installé. Ajoutez `openpyxl>=3.1` à requirements.txt."
            ) from exc
        wb = load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None) or []
        headers = [_normalize(h) for h in headers]
        for raw in rows:
            if not raw or all(v is None or str(v).strip() == "" for v in raw):
                continue
            yield {
                COLUMN_ALIASES.get(headers[i], headers[i]): raw[i]
                for i in range(len(headers))
                if i < len(raw)
            }
    elif name.endswith(".csv"):
        text = io.TextIOWrapper(file_obj, encoding="utf-8-sig", newline="")
        reader = csv.DictReader(text)
        norm_field = {f: COLUMN_ALIASES.get(_normalize(f), _normalize(f)) for f in reader.fieldnames or []}
        for raw in reader:
            yield {norm_field[k]: v for k, v in raw.items() if k in norm_field}
    else:
        raise ValueError("Format de fichier non supporté. Utilisez .xlsx ou .csv.")


def _coerce_field(field_name: str, value):
    """Normalise une valeur selon le champ cible."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return ""
    if field_name == "marital_status":
        return MARITAL_ALIASES.get(_normalize(value), str(value).strip().lower())
    if field_name == "gender":
        return GENDER_ALIASES.get(_normalize(value), str(value).strip().lower())
    if field_name == "dependents_count":
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0
    if field_name == "birth_date":
        import datetime
        if isinstance(value, datetime.date):
            return value
        s = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None
    if field_name == "last_name":
        return str(value).strip().upper()
    if field_name == "first_name":
        return " ".join(part.capitalize() for part in str(value).strip().split())
    if field_name == "email":
        return str(value).strip().lower()
    return str(value).strip()


def parse_members_file(file_obj, filename: str, mutuelle) -> ImportReport:
    """Importe en masse des membres d'une mutuelle à partir d'un fichier.

    Args:
        file_obj: file-like (UploadedFile, BytesIO).
        filename: nom du fichier (pour détecter .xlsx / .csv).
        mutuelle: instance Mutuelle d'affectation.

    Returns:
        ImportReport avec membres créés et erreurs détaillées.
    """
    report = ImportReport()

    try:
        rows = list(_read_rows(file_obj, filename))
    except (ValueError, RuntimeError) as exc:
        report.errors.append(ImportError(line=0, field="file", message=str(exc)))
        return report

    from django.db import transaction

    for line_no, row in enumerate(rows, start=2):  # 1 = header
        cleaned = {}
        # Coercition + check des champs requis
        for field_name, raw_value in row.items():
            if field_name in {None, ""}:
                continue
            cleaned[field_name] = _coerce_field(field_name, raw_value)
        missing = [f for f in REQUIRED_FIELDS if not cleaned.get(f)]
        if missing:
            for f in missing:
                report.errors.append(ImportError(line=line_no, field=f, message="Champ requis manquant"))
            continue

        # Doublon (téléphone OU email) dans la mutuelle ?
        dup_q = Member.all_objects.filter(mutuelle=mutuelle, phone=cleaned["phone"])
        if cleaned.get("email"):
            dup_q = dup_q | Member.all_objects.filter(
                mutuelle=mutuelle, email=cleaned["email"]
            )
        if dup_q.exists():
            report.skipped_duplicates += 1
            continue

        # Code membre auto si absent
        member_code = cleaned.pop("member_code", "") or _generate_member_code(mutuelle)
        qr_token = f"qr-{member_code.lower()}-{uuid.uuid4().hex[:8]}"

        try:
            with transaction.atomic():
                member = Member.all_objects.create(
                    mutuelle=mutuelle,
                    member_code=member_code,
                    qr_token=qr_token,
                    joined_at=timezone.now().date(),
                    **{k: v for k, v in cleaned.items() if v not in (None, "")},
                )
            report.created.append(member)
        except Exception as exc:  # noqa: BLE001 — capture pour rapport ligne
            report.errors.append(ImportError(line=line_no, field="-", message=str(exc)))

    return report


def _generate_member_code(mutuelle) -> str:
    slug_prefix = "".join(c for c in mutuelle.slug.upper() if c.isalnum())[:4] or "MEMB"
    year = timezone.now().year
    base = f"{slug_prefix}-{year}"
    seq = Member.all_objects.filter(mutuelle=mutuelle, member_code__startswith=base).count() + 1
    while True:
        candidate = f"{base}-{seq:04d}"
        if not Member.all_objects.filter(mutuelle=mutuelle, member_code=candidate).exists():
            return candidate
        seq += 1


# --- Invitations -----------------------------------------------------------
def send_member_invitation(invitation: MemberInvitation, request=None) -> bool:
    """Envoie le mail d'invitation pour rejoindre la mutuelle."""
    url = invitation.absolute_accept_url(request=request)
    ctx = {
        "invitation": invitation,
        "mutuelle": invitation.mutuelle,
        "accept_url": url,
    }
    subject = f"[{invitation.mutuelle.name}] Invitation à rejoindre la mutuelle"
    body_text = render_to_string("emails/member_invitation.txt", ctx)
    body_html = render_to_string("emails/member_invitation.html", ctx)
    from_email = getattr(settings, "DEFAULT_FROM_EMAIL", None) or f"no-reply@{invitation.mutuelle.slug}.mutuellex.com"
    try:
        send_mail(
            subject=subject,
            message=body_text,
            html_message=body_html,
            from_email=from_email,
            recipient_list=[invitation.email],
            fail_silently=False,
        )
        invitation.status = MemberInvitation.Status.SENT
        invitation.sent_at = timezone.now()
        invitation.save(update_fields=["status", "sent_at"])
        return True
    except Exception:
        # En dev/local sans SMTP : on garde l'invitation en PENDING.
        # Le lien est accessible côté admin ou via le retour de la vue.
        return False


def create_invitation(
    mutuelle, email: str, *, invited_by=None, full_name: str = "",
    message: str = "", ttl_days: int = 14,
) -> MemberInvitation:
    return MemberInvitation.all_objects.create(
        mutuelle=mutuelle,
        email=email.strip().lower(),
        full_name=full_name.strip(),
        invited_by=invited_by,
        expires_at=timezone.now() + timedelta(days=ttl_days),
        message=message.strip(),
    )
